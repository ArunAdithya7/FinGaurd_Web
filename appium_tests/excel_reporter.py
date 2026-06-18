"""
excel_reporter.py
-----------------
Generates a rich, multi-sheet Excel report for the FinGuard Appium E2E test run.

Sheets produced:
  1. Dashboard  — Summary banner, execution metadata, pass/fail statistics
  2. Detailed Logs — Full step table with status colouring & screenshot hyperlinks
  3. Analysis Charts — Data tables for pass-rate bar chart & duration chart (requires openpyxl chart engine)

Dependencies:
    pip install openpyxl>=3.1.2
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel


# ----------------------------------------------------------------
# Colour Palette — FinGuard Blue/Navy Premium Theme
# ----------------------------------------------------------------
C_NAVY_DARK   = "08142F"
C_BLUE_ACCENT = "2457F5"
C_LIGHT_BLUE  = "F4F7FB"
C_WHITE       = "FFFFFF"
C_BORDER      = "E5EAF2"

C_GREEN_BG    = "D1FAE5"
C_GREEN_TEXT  = "065F46"
C_RED_BG      = "FEE2E2"
C_RED_TEXT    = "991B1B"
C_YELLOW_BG   = "FEF9C3"
C_YELLOW_TEXT = "78350F"
C_GRAY_BG     = "F1F5F9"
C_GRAY_TEXT   = "475569"


def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name="Segoe UI", size=size, bold=bold, color=color, italic=italic)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ================================================================
# ExcelReporter
# ================================================================
class ExcelReporter:

    def __init__(self, report_dir="reports", screenshot_dir="screenshots"):
        self.report_dir     = report_dir
        self.screenshot_dir = screenshot_dir
        os.makedirs(self.report_dir,     exist_ok=True)
        os.makedirs(self.screenshot_dir, exist_ok=True)

    # ----------------------------------------------------------------
    # Public entry point
    # ----------------------------------------------------------------
    def generate_report(self, steps_data: list[dict],
                        start_time: datetime, end_time: datetime) -> str:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        # ---- pre-compute metrics ----
        total   = len(steps_data)
        passed  = sum(1 for s in steps_data if s["status"] == "PASS")
        failed  = sum(1 for s in steps_data if s["status"] == "FAIL")
        skipped = sum(1 for s in steps_data if s["status"] == "SKIP")
        pass_pct = round(passed / total * 100, 1) if total else 0.0
        dur_sec  = (end_time - start_time).total_seconds()
        dur_str  = f"{int(dur_sec // 60)}m {int(dur_sec % 60)}s"

        self._build_dashboard(wb, steps_data, start_time, end_time,
                              total, passed, failed, skipped, pass_pct, dur_str)
        self._build_detailed_logs(wb, steps_data, total, passed, failed, skipped)
        self._build_analysis_chart(wb, steps_data)

        filename = f"FinGuard_E2E_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        wb.save(filepath)
        return filepath

    # ================================================================
    # SHEET 1 — Dashboard
    # ================================================================
    def _build_dashboard(self, wb, steps_data, start_time, end_time,
                         total, passed, failed, skipped, pass_pct, dur_str):
        ws = wb.create_sheet(title="Dashboard")
        ws.sheet_view.showGridLines = False

        # ---- Banner ----
        ws.merge_cells("A1:G3")
        banner = ws["A1"]
        banner.value     = "🛡️  FinGuard — Appium E2E Automation Report"
        banner.font      = _font(size=20, bold=True, color=C_WHITE)
        banner.fill      = _fill(C_NAVY_DARK)
        banner.alignment = _align(h="center", v="center")
        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 10

        # ---- Sub-header row ----
        ws.merge_cells("A4:G4")
        sub = ws["A4"]
        sub.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
                         f"Platform: Android / UiAutomator2")
        sub.font      = _font(size=10, italic=True, color=C_GRAY_TEXT)
        sub.fill      = _fill(C_LIGHT_BLUE)
        sub.alignment = _align(h="center", v="center")
        ws.row_dimensions[4].height = 22

        ws.row_dimensions[5].height = 10  # spacer

        # ---- Section: Execution Metadata ----
        self._section_header(ws, row=6, col=1, text="⏱  Execution Metadata", span=3)
        meta_rows = [
            ("Start Time",         start_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("End Time",           end_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Duration",     dur_str),
            ("Platform",           "Android"),
            ("Automation Engine",  "UiAutomator2 (Appium)"),
            ("App Under Test",     "com.example.finguard"),
        ]
        for i, (k, v) in enumerate(meta_rows, start=7):
            ws.row_dimensions[i].height = 20
            c_k = ws.cell(row=i, column=1, value=k)
            c_k.font = _font(bold=True); c_k.border = _border()
            c_k.alignment = _align()

            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
            c_v = ws.cell(row=i, column=2, value=v)
            c_v.font = _font(); c_v.border = _border()
            c_v.alignment = _align()

        # ---- Section: Test Statistics ----
        self._section_header(ws, row=6, col=5, text="📊  Execution Statistics", span=3)
        stats = [
            ("Total Test Steps",   total,            C_WHITE,      None),
            ("✅ Passed",          passed,           C_GREEN_TEXT, C_GREEN_BG),
            ("❌ Failed",          failed,           C_RED_TEXT,   C_RED_BG),
            ("⏭️ Skipped",         skipped,          C_YELLOW_TEXT,C_YELLOW_BG),
            ("Pass Rate",          f"{pass_pct:.1f}%",
             C_GREEN_TEXT if pass_pct >= 80 else (C_RED_TEXT if pass_pct < 50 else C_YELLOW_TEXT),
             C_GREEN_BG   if pass_pct >= 80 else (C_RED_BG   if pass_pct < 50 else C_YELLOW_BG)),
        ]
        for i, (label, value, txt_c, bg_c) in enumerate(stats, start=7):
            ws.row_dimensions[i].height = 20
            c_l = ws.cell(row=i, column=5, value=label)
            c_l.font = _font(bold=True); c_l.border = _border(); c_l.alignment = _align()

            ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)
            c_v = ws.cell(row=i, column=6, value=value)
            c_v.font      = _font(bold=True, color=txt_c)
            c_v.border    = _border()
            c_v.alignment = _align(h="center")
            if bg_c:
                c_v.fill = _fill(bg_c)

        # ---- Status legend strip ----
        ws.row_dimensions[13].height = 10  # spacer
        self._section_header(ws, row=14, col=1, text="🗂  Test Case Summary", span=7)

        # mini table header
        mini_headers = ["#", "Test Case ID", "Test Name", "Status", "Duration (s)", "Timestamp"]
        mini_widths  = [4,   18,              40,          10,       14,              12]
        ws.row_dimensions[15].height = 22
        for ci, (h, _) in enumerate(zip(mini_headers, mini_widths), start=1):
            c = ws.cell(row=15, column=ci, value=h)
            c.font      = _font(bold=True, color=C_WHITE)
            c.fill      = _fill(C_BLUE_ACCENT)
            c.border    = _border()
            c.alignment = _align(h="center")

        # mini table data
        for ri, step in enumerate(steps_data, start=16):
            ws.row_dimensions[ri].height = 18
            data_row = [
                ri - 15,
                step["id"],
                step["name"],
                step["status"],
                f'{step["duration"]:.2f}',
                step["timestamp"].strftime("%H:%M:%S"),
            ]
            for ci, val in enumerate(data_row, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border    = _border()
                c.alignment = _align(h="center" if ci in (1, 4, 5, 6) else "left")
                # Zebra stripe
                c.fill = _fill(C_LIGHT_BLUE if (ri % 2 == 0) else C_WHITE)
                if ci == 4:  # Status column
                    if val == "PASS":
                        c.font = _font(bold=True, color=C_GREEN_TEXT); c.fill = _fill(C_GREEN_BG)
                    elif val == "FAIL":
                        c.font = _font(bold=True, color=C_RED_TEXT);   c.fill = _fill(C_RED_BG)
                    else:
                        c.font = _font(bold=True, color=C_YELLOW_TEXT); c.fill = _fill(C_YELLOW_BG)
                else:
                    c.font = _font()

        # ---- Column widths ----
        col_widths = {"A": 5, "B": 22, "C": 45, "D": 12, "E": 15, "F": 13, "G": 13}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

    # ================================================================
    # SHEET 2 — Detailed Logs
    # ================================================================
    def _build_detailed_logs(self, wb, steps_data, total, passed, failed, skipped):
        ws = wb.create_sheet(title="Detailed Logs")
        ws.sheet_view.showGridLines = True

        HEADERS = [
            "Step ID", "Step Name", "Status", "Duration (s)",
            "Timestamp", "Screenshot", "Error Details",
        ]
        COL_WIDTHS = [22, 42, 10, 14, 12, 20, 55]

        # Header row
        ws.row_dimensions[1].height = 28
        for ci, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = _font(size=11, bold=True, color=C_WHITE)
            c.fill      = _fill(C_NAVY_DARK)
            c.alignment = _align(h="center")
            c.border    = _border(color="000000", style="medium")

        # Data rows
        for ri, step in enumerate(steps_data, start=2):
            ws.row_dimensions[ri].height = 22
            row_fill = _fill(C_LIGHT_BLUE if ri % 2 == 0 else C_WHITE)

            # Step ID
            c = ws.cell(row=ri, column=1, value=step["id"])
            c.font = _font(bold=True); c.border = _border()
            c.alignment = _align(h="center"); c.fill = row_fill

            # Step Name
            c = ws.cell(row=ri, column=2, value=step["name"])
            c.font = _font(); c.border = _border()
            c.alignment = _align(); c.fill = row_fill

            # Status (conditional)
            c = ws.cell(row=ri, column=3, value=step["status"])
            c.alignment = _align(h="center"); c.border = _border()
            if step["status"] == "PASS":
                c.font = _font(bold=True, color=C_GREEN_TEXT); c.fill = _fill(C_GREEN_BG)
            elif step["status"] == "FAIL":
                c.font = _font(bold=True, color=C_RED_TEXT);   c.fill = _fill(C_RED_BG)
            else:
                c.font = _font(bold=True, color=C_YELLOW_TEXT); c.fill = _fill(C_YELLOW_BG)

            # Duration
            c = ws.cell(row=ri, column=4, value=step["duration"])
            c.number_format = "0.00"
            c.font = _font(); c.border = _border()
            c.alignment = _align(h="right"); c.fill = row_fill

            # Timestamp
            c = ws.cell(row=ri, column=5,
                        value=step["timestamp"].strftime("%H:%M:%S"))
            c.font = _font(); c.border = _border()
            c.alignment = _align(h="center"); c.fill = row_fill

            # Screenshot hyperlink
            c = ws.cell(row=ri, column=6)
            c.border = _border(); c.alignment = _align(h="center"); c.fill = row_fill
            if step.get("screenshot"):
                scr_name = os.path.basename(step["screenshot"])
                c.value = f'=HYPERLINK("../screenshots/{scr_name}","📷 View")'
                c.font  = _font(color="0B8CFF")
            else:
                c.value = "—"; c.font = _font(color=C_GRAY_TEXT)

            # Error
            c = ws.cell(row=ri, column=7, value=step.get("error") or "")
            c.font = _font(size=9, color="880000")
            c.alignment = _align(wrap=True); c.border = _border(); c.fill = row_fill

        # Summary / totals row
        summary_row = len(steps_data) + 2
        ws.row_dimensions[summary_row].height = 24
        thick = Border(top=Side(style="medium", color="000000"),
                       bottom=Side(style="medium", color="000000"))

        for ci in range(1, 8):
            c = ws.cell(row=summary_row, column=ci)
            c.border = thick; c.fill = _fill(C_LIGHT_BLUE)
            c.font   = _font(bold=True)

        ws.cell(row=summary_row, column=1, value="TOTAL").alignment = _align(h="center")
        ws.cell(row=summary_row, column=2, value=f"{total} steps")
        ws.cell(row=summary_row, column=3,
                value=f"✅ {passed}  ❌ {failed}  ⏭️ {skipped}").alignment = _align(h="center")
        ws.cell(row=summary_row, column=4,
                value=f"=SUM(D2:D{summary_row - 1})").number_format = "0.00"
        ws.cell(row=summary_row, column=4).alignment = _align(h="right")

        # Column widths
        for ci, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:G{len(steps_data) + 1}"

    # ================================================================
    # SHEET 3 — Analysis Charts
    # ================================================================
    def _build_analysis_chart(self, wb, steps_data):
        ws = wb.create_sheet(title="Analysis & Charts")
        ws.sheet_view.showGridLines = False

        # ---- Title ----
        ws.merge_cells("A1:L2")
        t = ws["A1"]
        t.value     = "📈  FinGuard Test Run — Visual Analysis"
        t.font      = _font(size=16, bold=True, color=C_WHITE)
        t.fill      = _fill(C_NAVY_DARK)
        t.alignment = _align(h="center", v="center")
        ws.row_dimensions[1].height = 40

        # ---- Data table 1: Status breakdown ----
        self._section_header(ws, row=4, col=1, text="Result Breakdown", span=2)
        status_labels = ["PASS", "FAIL", "SKIP"]
        status_counts = [
            sum(1 for s in steps_data if s["status"] == lbl)
            for lbl in status_labels
        ]
        ws.cell(row=5, column=1, value="Status").font  = _font(bold=True)
        ws.cell(row=5, column=2, value="Count").font   = _font(bold=True)
        for i, (lbl, cnt) in enumerate(zip(status_labels, status_counts), start=6):
            ws.cell(row=i, column=1, value=lbl).font  = _font()
            ws.cell(row=i, column=2, value=cnt).font  = _font()

        # Bar chart for status
        chart1 = BarChart()
        chart1.type            = "col"
        chart1.title           = "Test Result Distribution"
        chart1.y_axis.title    = "Number of Tests"
        chart1.x_axis.title    = "Status"
        chart1.style           = 10
        chart1.width           = 14
        chart1.height          = 10
        chart1.grouping        = "clustered"
        chart1.gapWidth        = 80

        data1 = Reference(ws, min_col=2, min_row=5, max_row=8)
        cats1 = Reference(ws, min_col=1, min_row=6, max_row=8)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)

        # Custom series colours
        from openpyxl.drawing.fill import PatternFillProperties
        from openpyxl.chart.data_source import NumDataSource, NumRef
        try:
            s = chart1.series[0]
            pt_colors = [C_GREEN_BG[:-0], "D1FAE5", "FEE2E2", "FEF9C3"]
            for idx, hex_c in enumerate(["00B050", "FF0000", "FFC000"]):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties.solidFill = hex_c
                s.dPt.append(pt)
        except Exception:
            pass

        ws.add_chart(chart1, "D4")

        # ---- Data table 2: Duration per step ----
        self._section_header(ws, row=4, col=10, text="Step Durations (s)", span=2)
        ws.cell(row=5, column=10, value="Step").font    = _font(bold=True)
        ws.cell(row=5, column=11, value="Duration").font = _font(bold=True)
        for i, step in enumerate(steps_data, start=6):
            ws.cell(row=i, column=10, value=step["id"]).font     = _font(size=8)
            ws.cell(row=i, column=11, value=step["duration"]).font = _font(size=8)
            ws.cell(row=i, column=11).number_format = "0.00"

        # Bar chart for durations
        end_row = 5 + len(steps_data)
        chart2 = BarChart()
        chart2.type            = "bar"
        chart2.title           = "Duration per Test Step (seconds)"
        chart2.y_axis.title    = "Step ID"
        chart2.x_axis.title    = "Seconds"
        chart2.style           = 10
        chart2.width           = 20
        chart2.height          = max(10, len(steps_data) * 0.7)

        data2 = Reference(ws, min_col=11, min_row=5, max_row=end_row)
        cats2 = Reference(ws, min_col=10, min_row=6, max_row=end_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        try:
            chart2.series[0].graphicalProperties.solidFill = C_BLUE_ACCENT
        except Exception:
            pass

        ws.add_chart(chart2, "D20")

        # ---- Column widths ----
        for col, w in [("A", 10), ("B", 10), ("C", 2), ("J", 22), ("K", 12)]:
            ws.column_dimensions[col].width = w

    # ----------------------------------------------------------------
    # Utility: section header cell
    # ----------------------------------------------------------------
    def _section_header(self, ws, row, col, text, span=2):
        end_col = col + span - 1
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=end_col,
            )
        c = ws.cell(row=row, column=col, value=text)
        c.font      = _font(size=11, bold=True, color=C_NAVY_DARK)
        c.fill      = _fill(C_LIGHT_BLUE)
        c.alignment = _align()
        c.border    = Border(bottom=Side(style="medium", color=C_NAVY_DARK))
        ws.row_dimensions[row].height = 22


# ================================================================
# Standalone mock verification
# ================================================================
if __name__ == "__main__":
    from datetime import timedelta
    reporter = ExcelReporter(report_dir="mock_reports", screenshot_dir="mock_screenshots")

    start = datetime(2026, 6, 11, 9, 0, 0)
    end   = start + timedelta(minutes=6, seconds=42)

    mock_steps = [
        {"id": "TC_001_LAUNCH",              "name": "Verify Launch & Splash screen",                    "status": "PASS",  "duration": 4.21,  "timestamp": start,                          "screenshot": None, "error": None},
        {"id": "TC_002_SIGNUP",              "name": "Sign up a new user account",                       "status": "PASS",  "duration": 8.55,  "timestamp": start + timedelta(seconds=5),   "screenshot": None, "error": None},
        {"id": "TC_003_LOGIN_INVALID_CREDS", "name": "Login validation — wrong credentials",             "status": "PASS",  "duration": 3.10,  "timestamp": start + timedelta(seconds=15),  "screenshot": None, "error": None},
        {"id": "TC_004_LOGIN_EMPTY_FIELDS",  "name": "Login validation — empty fields",                  "status": "PASS",  "duration": 2.40,  "timestamp": start + timedelta(seconds=19),  "screenshot": None, "error": None},
        {"id": "TC_005_LOGIN_SUCCESS",       "name": "Successful login to Dashboard",                    "status": "PASS",  "duration": 5.80,  "timestamp": start + timedelta(seconds=22),  "screenshot": None, "error": None},
        {"id": "TC_006_DASHBOARD_CARDS",     "name": "Verify Dashboard financial summary cards",         "status": "PASS",  "duration": 3.00,  "timestamp": start + timedelta(seconds=29),  "screenshot": None, "error": None},
        {"id": "TC_007_QUICK_ACTIONS",       "name": "Verify Dashboard quick action tiles",              "status": "PASS",  "duration": 2.10,  "timestamp": start + timedelta(seconds=33),  "screenshot": None, "error": None},
        {"id": "TC_008_ADD_INCOME",          "name": "Add monthly income record",                        "status": "PASS",  "duration": 7.50,  "timestamp": start + timedelta(seconds=36),  "screenshot": None, "error": None},
        {"id": "TC_009_ADD_EXPENSE",         "name": "Add monthly expense record",                       "status": "PASS",  "duration": 7.10,  "timestamp": start + timedelta(seconds=45),  "screenshot": None, "error": None},
        {"id": "TC_010_ADD_DEBT",            "name": "Add debt record",                                  "status": "FAIL",  "duration": 6.30,  "timestamp": start + timedelta(seconds=53),  "screenshot": None, "error": "TimeoutException: Debt tab not found"},
        {"id": "TC_011_RISK_ANALYSIS",       "name": "Verify Risk Analysis screen",                      "status": "PASS",  "duration": 9.00,  "timestamp": start + timedelta(seconds=61),  "screenshot": None, "error": None},
        {"id": "TC_012_PREDICTIONS",         "name": "Verify Predictions/Forecast screen",               "status": "PASS",  "duration": 8.50,  "timestamp": start + timedelta(seconds=71),  "screenshot": None, "error": None},
        {"id": "TC_013_RECOMMENDATIONS",     "name": "Verify Recommendations via Drawer",                "status": "PASS",  "duration": 7.20,  "timestamp": start + timedelta(seconds=81),  "screenshot": None, "error": None},
        {"id": "TC_014_NAVIGATION",          "name": "Bottom navigation bar traversal",                  "status": "PASS",  "duration": 5.60,  "timestamp": start + timedelta(seconds=90),  "screenshot": None, "error": None},
        {"id": "TC_015_EDIT_PROFILE",        "name": "Edit profile dialog — update name",                "status": "PASS",  "duration": 4.80,  "timestamp": start + timedelta(seconds=97),  "screenshot": None, "error": None},
        {"id": "TC_016_LOGOUT",              "name": "Logout from Profile and return to Login",          "status": "PASS",  "duration": 3.20,  "timestamp": start + timedelta(seconds=103), "screenshot": None, "error": None},
    ]

    saved = reporter.generate_report(mock_steps, start, end)
    print(f"\n✅  Mock Excel report saved to:\n    {saved}")
