# -*- coding: utf-8 -*-
"""
categorize_report.py
---------------------
Processes a generated FinGuard Appium E2E Excel report, extracts all passed test cases,
categorizes them into 11 core testing attributes, sorts them, and exports them into
a new, professionally styled Excel report.

Attributes:
  1. Functional Testing
  2. UI/UX Testing
  3. Compatibility Testing
  4. Performance Testing
  5. Security Testing
  6. API Testing
  7. Database Testing
  8. Accessibility Testing
  9. Mobile-Specific Testing
  10. Regression Testing
  11. End-to-End (E2E) Testing
"""

import os
import io
import sys
import glob
from datetime import datetime

# Force UTF-8 output on Windows to support Unicode characters in print()
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------
# Premium FinGuard Navy/Blue Styling Theme
# ----------------------------------------------------------------
C_NAVY_DARK   = "08142F"
C_BLUE_ACCENT = "2457F5"
C_LIGHT_BLUE  = "F4F7FB"
C_WHITE       = "FFFFFF"
C_BORDER      = "E5EAF2"

C_GREEN_BG    = "D1FAE5"
C_GREEN_TEXT  = "065F46"
C_GRAY_TEXT   = "475569"


def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name="Segoe UI", size=size, bold=bold, color=color, italic=italic)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ================================================================
# CATEGORIES MAPPING
# ================================================================
CATEGORIES_MAPPING = {
    "Functional Testing": [
        "TC_001", "TC_002", "TC_003", "TC_004", "TC_005", "TC_006", "TC_007", "TC_008", "TC_009", "TC_010",
        "TC_011", "TC_012", "TC_013", "TC_014", "TC_015", "TC_016", "TC_017", "TC_018", "TC_020", "TC_021",
        "TC_022", "TC_023", "TC_024", "TC_025", "TC_026", "TC_027", "TC_028", "TC_029", "TC_030", "TC_031",
        "TC_032", "TC_033", "TC_034", "TC_035", "TC_038", "TC_039", "TC_040", "TC_041", "TC_042", "TC_043",
        "TC_044", "TC_046", "TC_047", "TC_048", "TC_049", "TC_050", "TC_051", "TC_052", "TC_053", "TC_054",
        "TC_055", "TC_056", "TC_057", "TC_058", "TC_059", "TC_060", "TC_061", "TC_062", "TC_063", "TC_064",
        "TC_065", "TC_066", "TC_067", "TC_068", "TC_069", "TC_070", "TC_071", "TC_072", "TC_073", "TC_074",
        "TC_075", "TC_076", "TC_077", "TC_078", "TC_079", "TC_080", "TC_081", "TC_082", "TC_083", "TC_084",
        "TC_085", "TC_086", "TC_087", "TC_088", "TC_089", "TC_090", "TC_091", "TC_092", "TC_093", "TC_094",
        "TC_095", "TC_098", "TC_099", "TC_100", "TC_101", "TC_105", "TC_106", "TC_107", "TC_111", "TC_112",
        "TC_113", "TC_118", "TC_119", "TC_120", "TC_121", "TC_123"
    ],
    "UI/UX Testing": [
        "TC_001", "TC_004", "TC_019", "TC_021", "TC_036", "TC_037", "TC_038", "TC_039", "TC_040", "TC_041",
        "TC_042", "TC_043", "TC_044", "TC_045", "TC_046", "TC_047", "TC_048", "TC_049", "TC_050", "TC_051",
        "TC_066", "TC_081", "TC_096", "TC_097", "TC_098", "TC_099", "TC_100", "TC_101", "TC_102", "TC_103",
        "TC_104", "TC_105", "TC_106", "TC_108", "TC_109", "TC_110", "TC_111", "TC_112", "TC_113", "TC_115",
        "TC_117", "TC_118", "TC_121"
    ],
    "Compatibility Testing": [
        "TC_015", "TC_016", "TC_017", "TC_029", "TC_030", "TC_031", "TC_032", "TC_033", "TC_034", "TC_058",
        "TC_059", "TC_060", "TC_073", "TC_074", "TC_075", "TC_093", "TC_115", "TC_116"
    ],
    "Performance Testing": [
        "TC_001", "TC_020", "TC_035", "TC_036", "TC_045", "TC_061", "TC_062", "TC_063", "TC_064", "TC_065",
        "TC_076", "TC_077", "TC_078", "TC_079", "TC_080", "TC_094", "TC_095", "TC_097", "TC_102", "TC_106",
        "TC_108", "TC_120"
    ],
    "Security Testing": [
        "TC_006", "TC_007", "TC_008", "TC_009", "TC_014", "TC_018", "TC_022", "TC_025", "TC_026", "TC_028",
        "TC_029", "TC_030", "TC_031", "TC_032", "TC_119", "TC_123"
    ],
    "API Testing": [
        "TC_018", "TC_020", "TC_035", "TC_061", "TC_062", "TC_063", "TC_064", "TC_065", "TC_076", "TC_077",
        "TC_078", "TC_079", "TC_080", "TC_094", "TC_095", "TC_098", "TC_099", "TC_100", "TC_101", "TC_105",
        "TC_106", "TC_107", "TC_120", "TC_123"
    ],
    "Database Testing": [
        "TC_018", "TC_020", "TC_035", "TC_038", "TC_039", "TC_040", "TC_041", "TC_042", "TC_043", "TC_061",
        "TC_062", "TC_063", "TC_064", "TC_065", "TC_076", "TC_077", "TC_078", "TC_079", "TC_080", "TC_094",
        "TC_095", "TC_116", "TC_120"
    ],
    "Accessibility Testing": [
        "TC_001", "TC_021", "TC_037", "TC_038", "TC_039", "TC_040", "TC_041", "TC_042", "TC_043", "TC_044",
        "TC_051", "TC_066", "TC_081", "TC_097", "TC_104", "TC_109", "TC_110", "TC_112", "TC_118"
    ],
    "Mobile-Specific Testing": [
        "TC_002", "TC_019", "TC_045", "TC_050", "TC_102", "TC_108", "TC_109", "TC_110", "TC_114", "TC_116",
        "TC_122"
    ],
    "Regression Testing": [
        "TC_002", "TC_019", "TC_050", "TC_114", "TC_116", "TC_121", "TC_122"
    ],
    "End-to-End (E2E) Testing": [
        "TC_003", "TC_020", "TC_035", "TC_061", "TC_062", "TC_063", "TC_064", "TC_065", "TC_076", "TC_077",
        "TC_078", "TC_079", "TC_080", "TC_094", "TC_095", "TC_098", "TC_105", "TC_111", "TC_115", "TC_120",
        "TC_123"
    ]
}


def find_latest_report(report_dir="reports"):
    """Find the latest generated report in the reports directory."""
    if not os.path.exists(report_dir):
        return None
    files = glob.glob(os.path.join(report_dir, "FinGuard_E2E_Report_*.xlsx"))
    if not files:
        return None
    # Sort files by modification time
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def extract_passed_test_cases(filepath):
    """
    Reads the 'Detailed Logs' sheet from the Excel report.
    Returns a list of dicts: [{'id': 'TC_001', 'name': 'Verify splash...', 'duration': 4.2}]
    for all test cases that have status PASS.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "Detailed Logs" not in wb.sheetnames:
        raise ValueError(f"Could not find 'Detailed Logs' sheet in report: {filepath}")

    ws = wb["Detailed Logs"]
    passed_cases = []

    # Columns:
    # 1. Step ID, 2. Step Name, 3. Status, 4. Duration (s), 5. Timestamp, 6. Screenshot, 7. Error Details
    # Header is on row 1, data starts from row 2.
    for row in range(2, ws.max_row + 1):
        step_id = ws.cell(row=row, column=1).value
        # Ignore empty rows or the summary row at the bottom
        if not step_id or step_id == "TOTAL":
            continue

        status = ws.cell(row=row, column=3).value
        if status == "PASS":
            name = ws.cell(row=row, column=2).value or ""
            dur = ws.cell(row=row, column=4).value or 0.0
            passed_cases.append({
                "id": str(step_id).strip(),
                "name": str(name).strip(),
                "duration": float(dur)
            })

    return passed_cases


def generate_categorized_report(passed_cases, source_filename, output_dir="reports"):
    """
    Generates a new workbook with test cases separated into columns by attribute
    and sorted by Test Case ID.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorized Passed Tests"
    ws.sheet_view.showGridLines = True

    # 1. Banner Title
    ws.merge_cells("A1:K3")
    banner = ws["A1"]
    banner.value     = "🛡️  FinGuard — Categorized Passed E2E Test Cases"
    banner.font      = _font(size=18, bold=True, color=C_WHITE)
    banner.fill      = _fill(C_NAVY_DARK)
    banner.alignment = _align(h="center", v="center")
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15

    # 2. Sub-header row
    ws.merge_cells("A4:K4")
    sub = ws["A4"]
    sub.value     = (f"Processed from source report: {os.path.basename(source_filename)}   |   "
                     f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sub.font      = _font(size=9, italic=True, color=C_GRAY_TEXT)
    sub.fill      = _fill(C_LIGHT_BLUE)
    sub.alignment = _align(h="center", v="center")
    ws.row_dimensions[4].height = 20

    # 3. Summary metrics widget
    ws.merge_cells("A6:K6")
    sect = ws["A6"]
    sect.value = "📊  Passed Test Cases Summary by Attribute"
    sect.font = _font(size=11, bold=True, color=C_NAVY_DARK)
    sect.fill = _fill(C_LIGHT_BLUE)
    sect.alignment = Alignment(horizontal="left", vertical="center")
    sect.border = Border(bottom=Side(style="medium", color=C_NAVY_DARK))
    ws.row_dimensions[6].height = 22

    # Map each passed test case to its matching categories
    categorized_data = {cat: [] for cat in CATEGORIES_MAPPING}
    
    for tc in passed_cases:
        tc_id = tc["id"]
        tc_name = tc["name"]
        
        # Check matching category list
        for category, prefixes in CATEGORIES_MAPPING.items():
            if any(tc_id.startswith(pref) for pref in prefixes):
                categorized_data[category].append(tc)

    # Sort each category alphabetically/numerically by test case ID
    for category in categorized_data:
        # Sort based on the integer representation of the ID if possible, else alphabetically
        def get_sort_key(x):
            parts = x["id"].split("_")
            if len(parts) > 1 and parts[1].isdigit():
                return int(parts[1])
            return x["id"]
        
        categorized_data[category].sort(key=get_sort_key)

    # Mini statistics list in row 8
    ws.row_dimensions[8].height = 20
    total_passed = len(passed_cases)
    
    # Let's write the counts row for each category
    # Columns A to K correspond to 11 categories in alphabetical/order requested
    categories_list = list(CATEGORIES_MAPPING.keys())
    
    # Table headers at row 11
    ws.row_dimensions[11].height = 28
    for col_idx, category in enumerate(categories_list, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Header Cell
        cell = ws.cell(row=11, column=col_idx, value=category)
        cell.font = _font(size=10, bold=True, color=C_WHITE)
        cell.fill = _fill(C_BLUE_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border(color="000000", style="medium")
        
        # Write the counts just above the table headers in row 9-10
        # Row 9: "Passed Count" Label
        # Row 10: Value
        cnt_label = ws.cell(row=8, column=col_idx, value=f"{len(categorized_data[category])} passed")
        cnt_label.font = _font(size=9, bold=True, color=C_GREEN_TEXT)
        cnt_label.alignment = _align(h="center", v="center")
        cnt_label.fill = _fill(C_GREEN_BG)
        cnt_label.border = _border()

    # Spacer row
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[7].height = 10
    ws.row_dimensions[9].height = 10
    ws.row_dimensions[10].height = 10

    # 4. Populate test cases under each category column
    # Find the maximum rows we need to write
    max_rows = max(len(categorized_data[cat]) for cat in categorized_data)
    
    for row_offset in range(max_rows):
        row_num = 12 + row_offset
        ws.row_dimensions[row_num].height = 36  # Give enough height to wrap names
        
        # Zebra stripe style based on data rows
        row_fill = _fill(C_LIGHT_BLUE if row_offset % 2 == 0 else C_WHITE)
        
        for col_idx, category in enumerate(categories_list, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = _border()
            cell.fill = row_fill
            cell.alignment = _align()
            
            list_of_tc = categorized_data[category]
            if row_offset < len(list_of_tc):
                tc = list_of_tc[row_offset]
                cell.value = f"{tc['id']}\n{tc['name']}"
                # Style test case ID in cell (unfortunately openpyxl doesn't support multiple fonts in a single cell easily
                # without rich text, so we'll just format the whole string nicely)
                cell.font = _font(size=9)
            else:
                cell.value = ""
                cell.font = _font(size=9)

    # Adjust column widths dynamically with a reasonable minimum/maximum limits
    for col_idx in range(1, 12):
        col_letter = get_column_letter(col_idx)
        # Category headers or cell values
        # Let's set a fixed width of 25 characters for all columns so it doesn't extend too wide
        # but is still readable when wrapped.
        ws.column_dimensions[col_letter].width = 28

    # Ensure output dir exists
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"FinGuard_Categorized_Report_{timestamp}.xlsx"
    output_filepath = os.path.join(output_dir, output_filename)
    wb.save(output_filepath)
    return output_filepath


def main():
    print("=" * 60)
    print("  📊  FinGuard Test Report Categorizer & Sorter")
    print("=" * 60)
    
    # Get report directory
    appium_dir = os.path.dirname(os.path.abspath(__file__))
    report_dir = os.path.join(appium_dir, "reports")
    
    # Allow user to specify file path or default to latest
    if len(sys.argv) > 1:
        target_file = sys.argv[1]
        if not os.path.isabs(target_file):
            target_file = os.path.join(appium_dir, target_file)
    else:
        target_file = find_latest_report(report_dir)
        
    if not target_file or not os.path.exists(target_file):
        print(f"[ERROR] No valid report file found at {target_file}")
        print("Please specify a valid Excel report file path or generate a report first.")
        sys.exit(1)
        
    print(f"📁 Reading latest report: {os.path.basename(target_file)}")
    
    try:
        passed_cases = extract_passed_test_cases(target_file)
        print(f"✅ Extracted {len(passed_cases)} PASSED test cases.")
        
        if not passed_cases:
            print("⚠️ No passed test cases found to categorize.")
            sys.exit(0)
            
        output_path = generate_categorized_report(passed_cases, target_file, report_dir)
        print(f"\n🎉 Successfully categorized and exported to:\n    {output_path}")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ Error occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
